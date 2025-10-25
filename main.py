import openpyxl
from openpyxl import Workbook
import random
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime, timedelta
import os

class DummyDataGenerator:
    def __init__(self, root):
        self.root = root
        self.root.title("ダミーデータ作成")
        self.root.geometry("1000x850")
        
        # 背景色を設定
        self.root.configure(bg='#f0f4f8')
        
        # 現在の年度を計算（4月始まり）
        today = datetime.now()
        if today.month >= 4:
            self.current_year = today.year
        else:
            self.current_year = today.year - 1
        
        # 重複チェック用のセット
        self.used_student_ids = set()
        self.used_emails = set()
        self.used_names = set()
        
        # 名前データベース
        self.setup_name_database()
        
        # スタイルを設定
        self.setup_style()
        self.create_widgets()
    
    def setup_name_database(self):
        """日本人の名前データベースを設定"""
        # 苗字データ（漢字、ひらがな）- 150種類
        self.last_names = [
            ("佐藤", "さとう"), ("鈴木", "すずき"), ("高橋", "たかはし"), ("田中", "たなか"),
            ("伊藤", "いとう"), ("渡辺", "わたなべ"), ("山本", "やまもと"), ("中村", "なかむら"),
            ("小林", "こばやし"), ("加藤", "かとう"), ("吉田", "よしだ"), ("山田", "やまだ"),
            ("佐々木", "ささき"), ("山口", "やまぐち"), ("松本", "まつもと"), ("井上", "いのうえ"),
            ("木村", "きむら"), ("林", "はやし"), ("斎藤", "さいとう"), ("清水", "しみず"),
            ("山崎", "やまざき"), ("森", "もり"), ("池田", "いけだ"), ("橋本", "はしもと"),
            ("阿部", "あべ"), ("石川", "いしかわ"), ("前田", "まえだ"), ("藤田", "ふじた"),
            ("後藤", "ごとう"), ("岡田", "おかだ"), ("長谷川", "はせがわ"), ("村上", "むらかみ"),
            ("近藤", "こんどう"), ("石井", "いしい"), ("坂本", "さかもと"), ("遠藤", "えんどう"),
            ("青木", "あおき"), ("藤井", "ふじい"), ("西村", "にしむら"), ("福田", "ふくだ"),
            ("太田", "おおた"), ("三浦", "みうら"), ("岡本", "おかもと"), ("藤原", "ふじわら"),
            ("松田", "まつだ"), ("中島", "なかじま"), ("小川", "おがわ"), ("中野", "なかの"),
            ("原田", "はらだ"), ("竹内", "たけうち"), ("田村", "たむら"), ("金子", "かねこ"),
            ("和田", "わだ"), ("中山", "なかやま"), ("石田", "いしだ"), ("上田", "うえだ"),
            ("森田", "もりた"), ("原", "はら"), ("柴田", "しばた"), ("酒井", "さかい"),
            ("宮崎", "みやざき"), ("横山", "よこやま"), ("宮本", "みやもと"), ("内田", "うちだ"),
            ("高木", "たかぎ"), ("安藤", "あんどう"), ("谷口", "たにぐち"), ("大野", "おおの"),
            ("丸山", "まるやま"), ("今井", "いまい"), ("河野", "こうの"), ("藤本", "ふじもと"),
            ("村田", "むらた"), ("武田", "たけだ"), ("上野", "うえの"), ("杉山", "すぎやま"),
            ("千葉", "ちば"), ("増田", "ますだ"), ("小野", "おの"), ("久保", "くぼ"),
            ("市川", "いちかわ"), ("野口", "のぐち"), ("榊原", "さかきばら"), ("菊地", "きくち"),
            ("新井", "あらい"), ("大塚", "おおつか"), ("古川", "ふるかわ"), ("小島", "こじま"),
            ("水野", "みずの"), ("平野", "ひらの"), ("桜井", "さくらい"), ("木下", "きのした"),
            ("野村", "のむら"), ("松井", "まつい"), ("菅原", "すがわら"), ("高田", "たかだ"),
            ("北村", "きたむら"), ("大西", "おおにし"), ("小山", "こやま"), ("島田", "しまだ"),
            ("沢田", "さわだ"), ("工藤", "くどう"), ("西田", "にしだ"), ("服部", "はっとり"),
            ("樋口", "ひぐち"), ("内藤", "ないとう"), ("五十嵐", "いがらし"), ("浜田", "はまだ"),
            ("川口", "かわぐち"), ("関", "せき"), ("吉川", "よしかわ"), ("須藤", "すどう"),
            ("永井", "ながい"), ("岩崎", "いわさき"), ("田口", "たぐち"), ("本田", "ほんだ"),
            ("川上", "かわかみ"), ("杉本", "すぎもと"), ("中川", "なかがわ"), ("松尾", "まつお"),
            ("西川", "にしかわ"), ("大橋", "おおはし"), ("平田", "ひらた"), ("菊池", "きくち"),
            ("小松", "こまつ"), ("岩田", "いわた"), ("馬場", "ばば"), ("山下", "やました"),
            ("大島", "おおしま"), ("川崎", "かわさき"), ("矢野", "やの"), ("佐野", "さの"),
            ("松岡", "まつおか"), ("片山", "かたやま"), ("三宅", "みやけ"), ("飯田", "いいだ"),
            ("野田", "のだ"), ("中西", "なかにし"), ("吉岡", "よしおか"), ("山内", "やまうち"),
            ("黒田", "くろだ"), ("尾崎", "おざき"), ("辻", "つじ"), ("松村", "まつむら")
        ]
        
        # 男の名前データ（漢字、ひらがな）- 100種類
        self.male_first_names = [
            ("太郎", "たろう"), ("一郎", "いちろう"), ("二郎", "じろう"), ("三郎", "さぶろう"),
            ("健", "けん"), ("誠", "まこと"), ("隆", "たかし"), ("修", "おさむ"),
            ("勇", "いさむ"), ("翔", "しょう"), ("拓海", "たくみ"), ("大輝", "だいき"),
            ("翔太", "しょうた"), ("陽介", "ようすけ"), ("健太", "けんた"), ("雄大", "ゆうだい"),
            ("直樹", "なおき"), ("和也", "かずや"), ("裕太", "ゆうた"), ("拓也", "たくや"),
            ("大樹", "だいき"), ("悠斗", "ゆうと"), ("颯太", "そうた"), ("陽翔", "はると"),
            ("蓮", "れん"), ("湊", "みなと"), ("大和", "やまと"), ("樹", "いつき"),
            ("颯", "はやて"), ("陸", "りく"), ("蒼", "あおい"), ("悠真", "ゆうま"),
            ("海斗", "かいと"), ("航", "わたる"), ("遼", "りょう"), ("奏太", "そうた"),
            ("蒼空", "そら"), ("悠人", "ゆうと"), ("陽向", "ひなた"), ("結翔", "ゆいと"),
            ("朝陽", "あさひ"), ("颯真", "そうま"), ("瑛太", "えいた"), ("蒼大", "そうた"),
            ("陽太", "ようた"), ("湊斗", "みなと"), ("悠生", "ゆうせい"), ("陽斗", "はると"),
            ("大地", "だいち"), ("健人", "けんと"), ("勇気", "ゆうき"), ("翔平", "しょうへい"),
            ("涼太", "りょうた"), ("優斗", "ゆうと"), ("智也", "ともや"), ("隼人", "はやと"),
            ("駿", "しゅん"), ("颯人", "はやと"), ("奏", "かなで"), ("碧", "あおい"),
            ("新", "あらた"), ("蒼真", "そうま"), ("湊太", "そうた"), ("陸斗", "りくと"),
            ("悠希", "ゆうき"), ("遥斗", "はるト"), ("颯汰", "そうた"), ("蓮斗", "れんと"),
            ("大雅", "たいが"), ("悠翔", "ゆうと"), ("陸人", "りくと"), ("蒼士", "そうし"),
            ("海翔", "かいと"), ("颯斗", "はやと"), ("湊人", "みなと"), ("悠馬", "ゆうま"),
            ("陽大", "ようた"), ("蒼汰", "そうた"), ("碧斗", "あおと"), ("遥人", "はると"),
            ("勇斗", "ゆうと"), ("颯希", "そうき"), ("蓮人", "れんと"), ("悠太", "ゆうた"),
            ("陽平", "ようへい"), ("大智", "だいち"), ("優太", "ゆうた"), ("海人", "かいと"),
            ("颯馬", "そうま"), ("蒼太", "そうた"), ("湊翔", "みなと"), ("悠斗", "ゆうと"),
            ("陽", "よう"), ("蓮太", "れんた"), ("碧人", "あおと"), ("遥太", "はるた"),
            ("勇人", "はやと"), ("颯太郎", "そうたろう"), ("蒼人", "あおと"), ("湊大", "そうた")
        ]
        
        # 女の名前データ（漢字、ひらがな）- 100種類
        self.female_first_names = [
            ("花子", "はなこ"), ("美子", "よしこ"), ("恵子", "けいこ"), ("洋子", "ようこ"),
            ("幸子", "さちこ"), ("明子", "あきこ"), ("由美", "ゆみ"), ("真理子", "まりこ"),
            ("順子", "じゅんこ"), ("さくら", "さくら"), ("美咲", "みさき"), ("優花", "ゆうか"),
            ("結衣", "ゆい"), ("陽菜", "ひな"), ("葵", "あおい"), ("凛", "りん"),
            ("美月", "みつき"), ("彩花", "あやか"), ("愛美", "まなみ"), ("七海", "ななみ"),
            ("莉子", "りこ"), ("心春", "こはる"), ("結月", "ゆづき"), ("咲良", "さくら"),
            ("芽依", "めい"), ("琴音", "ことね"), ("陽葵", "ひまり"), ("紬", "つむぎ"),
            ("楓", "かえで"), ("杏", "あん"), ("澪", "みお"), ("花音", "かのん"),
            ("結菜", "ゆいな"), ("心愛", "ここあ"), ("陽菜乃", "ひなの"), ("美羽", "みう"),
            ("心花", "このか"), ("彩乃", "あやの"), ("愛莉", "あいり"), ("結愛", "ゆあ"),
            ("陽向", "ひなた"), ("美桜", "みお"), ("心結", "みゆ"), ("花菜", "はな"),
            ("結心", "ゆみ"), ("陽咲", "ひさき"), ("美優", "みゆう"), ("愛花", "あいか"),
            ("心音", "ここね"), ("彩葉", "いろは"), ("結花", "ゆいか"), ("陽愛", "ひな"),
            ("美空", "みそら"), ("心菜", "ここな"), ("花凛", "かりん"), ("結奈", "ゆいな"),
            ("陽莉", "ひより"), ("美緒", "みお"), ("愛梨", "あいり"), ("心咲", "みさき"),
            ("彩音", "あやね"), ("結香", "ゆいか"), ("陽茉莉", "ひまり"), ("美波", "みなみ"),
            ("心優", "みゆう"), ("花梨", "かりん"), ("結彩", "ゆいあ"), ("陽花", "はるか"),
            ("美織", "みおり"), ("愛菜", "あいな"), ("心葉", "このは"), ("彩華", "あやか"),
            ("結莉", "ゆいり"), ("陽奈", "ひな"), ("美結", "みゆ"), ("心海", "ここみ"),
            ("花歩", "かほ"), ("結乃", "ゆいの"), ("陽菜子", "ひなこ"), ("美帆", "みほ"),
            ("愛実", "まなみ"), ("心彩", "こあ"), ("彩未", "あやみ"), ("結梨", "ゆいり"),
            ("陽依", "ひより"), ("美紗", "みさ"), ("心桜", "こころ"), ("花穂", "かほ"),
            ("結希", "ゆいき"), ("陽香", "はるか"), ("美咲希", "みさき"), ("愛佳", "あいか"),
            ("心陽", "こはる"), ("彩夏", "あやか"), ("結実", "ゆみ"), ("陽梨", "ひなり"),
            ("美琴", "みこと"), ("心愛美", "ここみ"), ("花乃", "かの"), ("結唯", "ゆい")
        ]
        
        # 保護者世代の名前データ（漢字、ひらがな）- 80種類
        self.guardian_first_names = [
            ("太郎", "たろう"), ("一郎", "いちろう"), ("二郎", "じろう"), ("健一", "けんいち"),
            ("誠", "まこと"), ("隆", "たかし"), ("修", "おさむ"), ("勇", "いさむ"),
            ("浩", "ひろし"), ("茂", "しげる"), ("清", "きよし"), ("正", "ただし"),
            ("実", "みのる"), ("進", "すすむ"), ("豊", "ゆたか"), ("昭", "あきら"),
            ("博", "ひろし"), ("明", "あきら"), ("武", "たけし"), ("剛", "つよし"),
            ("健", "けん"), ("強", "つよし"), ("勝", "まさる"), ("和夫", "かずお"),
            ("幸雄", "ゆきお"), ("正男", "まさお"), ("義雄", "よしお"), ("光男", "みつお"),
            ("秀雄", "ひでお"), ("利夫", "としお"), ("文雄", "ふみお"), ("勝彦", "かつひこ"),
            ("和彦", "かずひこ"), ("正彦", "まさひこ"), ("義彦", "よしひこ"), ("秀彦", "ひでひこ"),
            ("哲也", "てつや"), ("和也", "かずや"), ("雅也", "まさや"), ("裕也", "ゆうや"),
            ("花子", "はなこ"), ("美子", "よしこ"), ("恵子", "けいこ"), ("洋子", "ようこ"),
            ("幸子", "さちこ"), ("明子", "あきこ"), ("由美", "ゆみ"), ("真理子", "まりこ"),
            ("順子", "じゅんこ"), ("久美子", "くみこ"), ("直子", "なおこ"), ("裕子", "ゆうこ"),
            ("智子", "ともこ"), ("和子", "かずこ"), ("節子", "せつこ"), ("文子", "ふみこ"),
            ("春子", "はるこ"), ("夏子", "なつこ"), ("秋子", "あきこ"), ("冬子", "ふゆこ"),
            ("早苗", "さなえ"), ("理恵", "りえ"), ("美穂", "みほ"), ("麻美", "あさみ"),
            ("香織", "かおり"), ("愛", "あい"), ("舞", "まい"), ("綾", "あや"),
            ("美紀", "みき"), ("恵美", "えみ"), ("陽子", "ようこ"), ("加奈子", "かなこ"),
            ("千春", "ちはる"), ("美智子", "みちこ"), ("悦子", "えつこ"), ("律子", "りつこ"),
            ("雅子", "まさこ"), ("典子", "のりこ"), ("貴子", "たかこ"), ("敏子", "としこ")
        ]
    
    def setup_style(self):
        """カスタムスタイルを設定"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # LabelFrameのスタイル
        style.configure('Custom.TLabelframe', 
                       background='#ffffff',
                       borderwidth=2,
                       relief='solid')
        style.configure('Custom.TLabelframe.Label', 
                       background='#ffffff',
                       foreground='#2c3e50',
                       font=('Arial', 11, 'bold'))
        
        # Checkbuttonのスタイル - 大きくする
        style.configure('Custom.TCheckbutton',
                       background='#ffffff',
                       foreground='#2c3e50',
                       font=('Arial', 11))
    
    def create_widgets(self):
        # メインコンテナ
        container = tk.Frame(self.root, bg='#f0f4f8')
        container.pack(fill="both", expand=True)
        
        # キャンバスとスクロールバーを作成
        canvas = tk.Canvas(container, bg='#f0f4f8', highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f0f4f8')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # マウスホイールでスクロール
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # ファイル選択
        file_frame = ttk.LabelFrame(scrollable_frame, text="Excelファイル", 
                                    padding=15, style='Custom.TLabelframe')
        file_frame.pack(fill="x", padx=15, pady=8)
        
        self.file_path = tk.StringVar()
        file_entry = ttk.Entry(file_frame, textvariable=self.file_path, width=80, font=('Arial', 10))
        file_entry.pack(side="left", padx=5)
        
        browse_btn = tk.Button(file_frame, text="参照", command=self.select_file,
                              bg='#3498db', fg='white', font=('Arial', 10, 'bold'),
                              padx=15, pady=5, relief='flat', cursor='hand2',
                              activebackground='#2980b9', activeforeground='white')
        browse_btn.pack(side="left", padx=5)
        
        # シート名と共通設定を横並び
        settings_container = tk.Frame(scrollable_frame, bg='#f0f4f8')
        settings_container.pack(fill="x", padx=15, pady=8)
        
        # シート名
        sheet_frame = ttk.LabelFrame(settings_container, text="シート名", 
                                     padding=15, style='Custom.TLabelframe')
        sheet_frame.pack(side="left", fill="both", expand=True, padx=(0, 5))
        
        self.sheet_name = tk.StringVar(value="Sheet1")
        ttk.Entry(sheet_frame, textvariable=self.sheet_name, width=20, font=('Arial', 10)).pack()
        
        # 共通設定
        common_frame = ttk.LabelFrame(settings_container, text="共通設定", 
                                      padding=15, style='Custom.TLabelframe')
        common_frame.pack(side="left", fill="both", expand=True, padx=(5, 0))
        
        common_inner = tk.Frame(common_frame, bg='#ffffff')
        common_inner.pack()
        
        tk.Label(common_inner, text="開始行:", bg='#ffffff', fg='#2c3e50', 
                font=('Arial', 10)).grid(row=0, column=0, sticky="w", padx=5)
        self.start_row = tk.StringVar(value="2")
        ttk.Entry(common_inner, textvariable=self.start_row, width=10, 
                 font=('Arial', 10)).grid(row=0, column=1, padx=5)
        
        tk.Label(common_inner, text="件数:", bg='#ffffff', fg='#2c3e50', 
                font=('Arial', 10)).grid(row=0, column=2, sticky="w", padx=20)
        self.data_count = tk.StringVar(value="100")
        ttk.Entry(common_inner, textvariable=self.data_count, width=10, 
                 font=('Arial', 10)).grid(row=0, column=3, padx=5)
        
        # データ項目を2段レイアウトに
        # 左列
        left_column = tk.Frame(scrollable_frame, bg='#f0f4f8')
        left_column.pack(side="left", fill="both", expand=True, padx=(15, 7))
        
        # 右列
        right_column = tk.Frame(scrollable_frame, bg='#f0f4f8')
        right_column.pack(side="left", fill="both", expand=True, padx=(7, 15))
        
        # 学生番号設定（左列）
        self.student_id_enabled = tk.BooleanVar(value=False)
        student_id_frame = ttk.LabelFrame(left_column, 
                                          text="学生番号（2から始まる7桁の数字）", 
                                          padding=15, style='Custom.TLabelframe')
        student_id_frame.pack(fill="x", pady=8)
        
        student_id_inner = tk.Frame(student_id_frame, bg='#ffffff')
        student_id_inner.pack()
        
        student_id_check = tk.Checkbutton(student_id_inner, text="生成する", 
                                         variable=self.student_id_enabled,
                                         bg='#ffffff', fg='#2c3e50',
                                         font=('Arial', 11), selectcolor='#ffffff',
                                         activebackground='#ffffff',
                                         activeforeground='#3498db',
                                         cursor='hand2', padx=5, pady=5)
        student_id_check.grid(row=0, column=0, sticky="w", padx=5)
        
        tk.Label(student_id_inner, text="列名:", bg='#ffffff', fg='#2c3e50', 
                font=('Arial', 10)).grid(row=0, column=1, padx=20)
        self.student_id_column = tk.StringVar(value="A")
        ttk.Entry(student_id_inner, textvariable=self.student_id_column, 
                 width=10, font=('Arial', 10)).grid(row=0, column=2, padx=5)
        
        # 名前設定（左列）
        self.name_enabled = tk.BooleanVar(value=False)
        name_frame = ttk.LabelFrame(left_column, text="名前", 
                                    padding=15, style='Custom.TLabelframe')
        name_frame.pack(fill="x", pady=8)
        
        name_inner = tk.Frame(name_frame, bg='#ffffff')
        name_inner.pack()
        
        name_check = tk.Checkbutton(name_inner, text="生成する", 
                                   variable=self.name_enabled,
                                   bg='#ffffff', fg='#2c3e50',
                                   font=('Arial', 11), selectcolor='#ffffff',
                                   activebackground='#ffffff',
                                   activeforeground='#3498db',
                                   cursor='hand2', padx=5, pady=5)
        name_check.grid(row=0, column=0, sticky="w", padx=5)
        
        tk.Label(name_inner, text="列名:", bg='#ffffff', fg='#2c3e50', 
                font=('Arial', 10)).grid(row=0, column=1, padx=20)
        self.name_column = tk.StringVar(value="B")
        ttk.Entry(name_inner, textvariable=self.name_column, width=10, 
                 font=('Arial', 10)).grid(row=0, column=2, padx=5)
        
        # ふりがな設定（左列）
        self.furigana_enabled = tk.BooleanVar(value=False)
        furigana_frame = ttk.LabelFrame(left_column, text="ふりがな", 
                                        padding=15, style='Custom.TLabelframe')
        furigana_frame.pack(fill="x", pady=8)
        
        furigana_inner = tk.Frame(furigana_frame, bg='#ffffff')
        furigana_inner.pack()
        
        furigana_check = tk.Checkbutton(furigana_inner, text="生成する", 
                                       variable=self.furigana_enabled,
                                       bg='#ffffff', fg='#2c3e50',
                                       font=('Arial', 11), selectcolor='#ffffff',
                                       activebackground='#ffffff',
                                       activeforeground='#3498db',
                                       cursor='hand2', padx=5, pady=5)
        furigana_check.grid(row=0, column=0, sticky="w", padx=5)
        
        tk.Label(furigana_inner, text="列名:", bg='#ffffff', fg='#2c3e50', 
                font=('Arial', 10)).grid(row=0, column=1, padx=20)
        self.furigana_column = tk.StringVar(value="C")
        ttk.Entry(furigana_inner, textvariable=self.furigana_column, width=10, 
                 font=('Arial', 10)).grid(row=0, column=2, padx=5)
        
        # 性別設定（左列）
        self.gender_enabled = tk.BooleanVar(value=False)
        gender_frame = ttk.LabelFrame(left_column, text="性別", 
                                      padding=15, style='Custom.TLabelframe')
        gender_frame.pack(fill="x", pady=8)
        
        gender_inner = tk.Frame(gender_frame, bg='#ffffff')
        gender_inner.pack()
        
        gender_check = tk.Checkbutton(gender_inner, text="生成する", 
                                     variable=self.gender_enabled,
                                     bg='#ffffff', fg='#2c3e50',
                                     font=('Arial', 11), selectcolor='#ffffff',
                                     activebackground='#ffffff',
                                     activeforeground='#3498db',
                                     cursor='hand2', padx=5, pady=5)
        gender_check.grid(row=0, column=0, sticky="w", padx=5)
        
        tk.Label(gender_inner, text="列名:", bg='#ffffff', fg='#2c3e50', 
                font=('Arial', 10)).grid(row=0, column=1, padx=20)
        self.gender_column = tk.StringVar(value="D")
        ttk.Entry(gender_inner, textvariable=self.gender_column, width=10, 
                 font=('Arial', 10)).grid(row=0, column=2, padx=5)
        
        # 生年月日設定（左列）
        self.birthdate_enabled = tk.BooleanVar(value=False)
        birthdate_frame = ttk.LabelFrame(left_column, 
                                         text=f"生年月日（高校新入生 - {self.current_year}年度）", 
                                         padding=15, style='Custom.TLabelframe')
        birthdate_frame.pack(fill="x", pady=8)
        
        birthdate_inner = tk.Frame(birthdate_frame, bg='#ffffff')
        birthdate_inner.pack()
        
        birthdate_check = tk.Checkbutton(birthdate_inner, text="生成する", 
                                        variable=self.birthdate_enabled,
                                        bg='#ffffff', fg='#2c3e50',
                                        font=('Arial', 11), selectcolor='#ffffff',
                                        activebackground='#ffffff',
                                        activeforeground='#3498db',
                                        cursor='hand2', padx=5, pady=5)
        birthdate_check.grid(row=0, column=0, sticky="w", padx=5)
        
        tk.Label(birthdate_inner, text="列名:", bg='#ffffff', fg='#2c3e50', 
                font=('Arial', 10)).grid(row=0, column=1, padx=20)
        self.birthdate_column = tk.StringVar(value="E")
        ttk.Entry(birthdate_inner, textvariable=self.birthdate_column, 
                 width=10, font=('Arial', 10)).grid(row=0, column=2, padx=5)
        
        # 住所設定（右列）
        self.address_enabled = tk.BooleanVar(value=False)
        address_frame = ttk.LabelFrame(right_column, text="東京都の住所", 
                                       padding=15, style='Custom.TLabelframe')
        address_frame.pack(fill="x", pady=8)
        
        address_inner = tk.Frame(address_frame, bg='#ffffff')
        address_inner.pack()
        
        address_check = tk.Checkbutton(address_inner, text="生成する", 
                                      variable=self.address_enabled,
                                      bg='#ffffff', fg='#2c3e50',
                                      font=('Arial', 11), selectcolor='#ffffff',
                                      activebackground='#ffffff',
                                      activeforeground='#3498db',
                                      cursor='hand2', padx=5, pady=5)
        address_check.grid(row=0, column=0, sticky="w", padx=5)
        
        tk.Label(address_inner, text="列名:", bg='#ffffff', fg='#2c3e50', 
                font=('Arial', 10)).grid(row=0, column=1, padx=20)
        self.address_column = tk.StringVar(value="F")
        ttk.Entry(address_inner, textvariable=self.address_column, 
                 width=10, font=('Arial', 10)).grid(row=0, column=2, padx=5)
        
        # 保護者氏名設定（右列）
        self.guardian_name_enabled = tk.BooleanVar(value=False)
        guardian_name_frame = ttk.LabelFrame(right_column, 
                                             text="保護者氏名（学生と同じ苗字）", 
                                             padding=15, style='Custom.TLabelframe')
        guardian_name_frame.pack(fill="x", pady=8)
        
        guardian_name_inner = tk.Frame(guardian_name_frame, bg='#ffffff')
        guardian_name_inner.pack()
        
        guardian_name_check = tk.Checkbutton(guardian_name_inner, text="生成する", 
                                            variable=self.guardian_name_enabled,
                                            bg='#ffffff', fg='#2c3e50',
                                            font=('Arial', 11), selectcolor='#ffffff',
                                            activebackground='#ffffff',
                                            activeforeground='#3498db',
                                            cursor='hand2', padx=5, pady=5)
        guardian_name_check.grid(row=0, column=0, sticky="w", padx=5)
        
        tk.Label(guardian_name_inner, text="列名:", bg='#ffffff', fg='#2c3e50', 
                font=('Arial', 10)).grid(row=0, column=1, padx=20)
        self.guardian_name_column = tk.StringVar(value="G")
        ttk.Entry(guardian_name_inner, textvariable=self.guardian_name_column, 
                 width=10, font=('Arial', 10)).grid(row=0, column=2, padx=5)
        
        # 保護者ふりがな設定（右列）
        self.guardian_furigana_enabled = tk.BooleanVar(value=False)
        guardian_furigana_frame = ttk.LabelFrame(right_column, text="保護者ふりがな", 
                                                 padding=15, style='Custom.TLabelframe')
        guardian_furigana_frame.pack(fill="x", pady=8)
        
        guardian_furigana_inner = tk.Frame(guardian_furigana_frame, bg='#ffffff')
        guardian_furigana_inner.pack()
        
        guardian_furigana_check = tk.Checkbutton(guardian_furigana_inner, text="生成する", 
                                                variable=self.guardian_furigana_enabled,
                                                bg='#ffffff', fg='#2c3e50',
                                                font=('Arial', 11), selectcolor='#ffffff',
                                                activebackground='#ffffff',
                                                activeforeground='#3498db',
                                                cursor='hand2', padx=5, pady=5)
        guardian_furigana_check.grid(row=0, column=0, sticky="w", padx=5)
        
        tk.Label(guardian_furigana_inner, text="列名:", bg='#ffffff', fg='#2c3e50', 
                font=('Arial', 10)).grid(row=0, column=1, padx=20)
        self.guardian_furigana_column = tk.StringVar(value="H")
        ttk.Entry(guardian_furigana_inner, textvariable=self.guardian_furigana_column, 
                 width=10, font=('Arial', 10)).grid(row=0, column=2, padx=5)
        
        # 保護者電話番号設定（右列）
        self.guardian_phone_enabled = tk.BooleanVar(value=False)
        guardian_phone_frame = ttk.LabelFrame(right_column, 
                                              text="保護者電話番号", 
                                              padding=15, style='Custom.TLabelframe')
        guardian_phone_frame.pack(fill="x", pady=8)
        
        guardian_phone_inner = tk.Frame(guardian_phone_frame, bg='#ffffff')
        guardian_phone_inner.pack()
        
        guardian_phone_check = tk.Checkbutton(guardian_phone_inner, text="生成する", 
                                             variable=self.guardian_phone_enabled,
                                             bg='#ffffff', fg='#2c3e50',
                                             font=('Arial', 11), selectcolor='#ffffff',
                                             activebackground='#ffffff',
                                             activeforeground='#3498db',
                                             cursor='hand2', padx=5, pady=5)
        guardian_phone_check.grid(row=0, column=0, sticky="w", padx=5)
        
        tk.Label(guardian_phone_inner, text="列名:", bg='#ffffff', fg='#2c3e50', 
                font=('Arial', 10)).grid(row=0, column=1, padx=20)
        self.guardian_phone_column = tk.StringVar(value="I")
        ttk.Entry(guardian_phone_inner, textvariable=self.guardian_phone_column, 
                 width=10, font=('Arial', 10)).grid(row=0, column=2, padx=5)
        
        # メールアドレス設定（右列）
        self.email_enabled = tk.BooleanVar(value=False)
        email_frame = ttk.LabelFrame(right_column, 
                                     text="メールアドレス（2から始まる10桁@ドメイン）", 
                                     padding=15, style='Custom.TLabelframe')
        email_frame.pack(fill="x", pady=8)
        
        email_inner = tk.Frame(email_frame, bg='#ffffff')
        email_inner.pack()
        
        email_check = tk.Checkbutton(email_inner, text="生成する", 
                                    variable=self.email_enabled,
                                    bg='#ffffff', fg='#2c3e50',
                                    font=('Arial', 11), selectcolor='#ffffff',
                                    activebackground='#ffffff',
                                    activeforeground='#3498db',
                                    cursor='hand2', padx=5, pady=5)
        email_check.grid(row=0, column=0, sticky="w", padx=5)
        
        tk.Label(email_inner, text="列名:", bg='#ffffff', fg='#2c3e50', 
                font=('Arial', 10)).grid(row=0, column=1, padx=20)
        self.email_column = tk.StringVar(value="J")
        ttk.Entry(email_inner, textvariable=self.email_column, 
                 width=10, font=('Arial', 10)).grid(row=0, column=2, padx=5)
        
        # 実行ボタン（両列の下に配置）
        button_container = tk.Frame(scrollable_frame, bg='#f0f4f8')
        button_container.pack(fill="x", pady=25, padx=15)
        
        generate_btn = tk.Button(button_container, 
                                text="データ生成", 
                                command=self.generate_data,
                                bg='#e74c3c',
                                fg='white',
                                font=('Arial', 16, 'bold'),
                                padx=50,
                                pady=15,
                                relief='flat',
                                cursor='hand2',
                                activebackground='#c0392b',
                                activeforeground='white',
                                borderwidth=0)
        generate_btn.pack()
        
        # ホバー効果
        def on_enter(e):
            generate_btn['background'] = '#c0392b'
        def on_leave(e):
            generate_btn['background'] = '#e74c3c'
        
        generate_btn.bind("<Enter>", on_enter)
        generate_btn.bind("<Leave>", on_leave)
        
        # キャンバスとスクロールバーを配置
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
    def select_file(self):
        filename = filedialog.asksaveasfilename(
            title="Excelファイルを選択または作成",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.file_path.set(filename)
    
    def generate_student_id(self):
        """2から始まる7桁の学生番号を生成（重複なし）"""
        max_attempts = 10000
        for _ in range(max_attempts):
            student_id = f"2{random.randint(0, 999999):06d}"
            if student_id not in self.used_student_ids:
                self.used_student_ids.add(student_id)
                return student_id
        raise ValueError("学生番号の生成に失敗しました。件数を減らしてください。")
    
    def generate_name_with_furigana(self, is_male=None):
        """名前とふりがなを生成（性別指定可能、重複チェック付き）"""
        max_attempts = 1000
        
        for _ in range(max_attempts):
            last_name, last_name_kana = random.choice(self.last_names)
            
            if is_male is None:
                is_male = random.choice([True, False])
            
            if is_male:
                first_name, first_name_kana = random.choice(self.male_first_names)
            else:
                first_name, first_name_kana = random.choice(self.female_first_names)
            
            full_name = f"{last_name} {first_name}"
            
            # 重複チェック
            if full_name not in self.used_names:
                self.used_names.add(full_name)
                full_name_kana = f"{last_name_kana} {first_name_kana}"
                return full_name, full_name_kana, last_name, last_name_kana
        
        # 重複が多すぎる場合は警告を出さずに重複を許可
        full_name_kana = f"{last_name_kana} {first_name_kana}"
        return full_name, full_name_kana, last_name, last_name_kana
    
    def generate_tokyo_address(self):
        """東京都の住所を生成"""
        tokyo_wards = [
            "千代田区", "中央区", "港区", "新宿区", "文京区", "台東区", "墨田区",
            "江東区", "品川区", "目黒区", "大田区", "世田谷区", "渋谷区", "中野区",
            "杉並区", "豊島区", "北区", "荒川区", "板橋区", "練馬区", "足立区", "葛飾区", "江戸川区"
        ]
        
        town_names = [
            "青山", "赤坂", "麻布", "市ヶ谷", "神楽坂", "九段", "麹町", "四谷",
            "銀座", "日本橋", "八重洲", "築地", "月島", "晴海",
            "六本木", "白金", "高輪", "芝", "浜松町", "台場",
            "歌舞伎町", "西新宿", "早稲田", "高田馬場", "神楽坂",
            "本郷", "湯島", "根津", "千駄木", "白山",
            "上野", "浅草", "蔵前", "浅草橋",
            "両国", "錦糸町", "押上", "向島",
            "豊洲", "東雲", "有明", "辰巳",
            "五反田", "大崎", "戸越", "荏原",
            "自由が丘", "中目黒", "祐天寺", "学芸大学",
            "蒲田", "大森", "田園調布", "雪谷",
            "三軒茶屋", "下北沢", "経堂", "成城",
            "渋谷", "恵比寿", "代官山", "原宿",
            "中野", "東中野", "野方", "鷺ノ宮",
            "高円寺", "阿佐ヶ谷", "荻窪", "西荻窪",
            "池袋", "巣鴨", "駒込", "大塚",
            "赤羽", "王子", "田端", "西ヶ原",
            "日暮里", "町屋", "西日暮里",
            "板橋", "大山", "成増", "志村",
            "練馬", "石神井", "大泉", "光が丘",
            "北千住", "西新井", "竹ノ塚", "綾瀬",
            "亀有", "金町", "新小岩", "柴又",
            "小岩", "葛西", "西葛西", "船堀"
        ]
        
        ward = random.choice(tokyo_wards)
        town = random.choice(town_names)
        chome = random.randint(1, 9)
        banchi = random.randint(1, 30)
        go = random.randint(1, 20)
        
        return f"東京都{ward}{town}{chome}-{banchi}-{go}"
    
    def generate_birthdate(self):
        """高校新入生の生年月日を生成（同学年）"""
        birth_year = self.current_year - 15
        start_date = datetime(birth_year - 1, 4, 2)
        end_date = datetime(birth_year, 4, 1)
        
        days_between = (end_date - start_date).days
        random_days = random.randint(0, days_between)
        
        birth_date = start_date + timedelta(days=random_days)
        return birth_date.strftime("%Y/%m/%d")
    
    def generate_phone_number(self):
        """日本の電話番号を生成"""
        area_codes = ["03", "090", "080", "070"]
        area_code = random.choice(area_codes)
        return f"{area_code}-{random.randint(1000, 9999)}-{random.randint(1000, 9999)}"
    
    def generate_email(self):
        """2から始まる10桁の数字をユーザー名とするメールアドレスを生成（重複なし）"""
        max_attempts = 10000
        domains = ["example.com", "sample.co.jp", "test.jp", "dummy.com", "school.ac.jp"]
        
        for _ in range(max_attempts):
            user_part = f"2{random.randint(0, 999999999):09d}"
            domain = random.choice(domains)
            email = f"{user_part}@{domain}"
            
            if email not in self.used_emails:
                self.used_emails.add(email)
                return email
        
        raise ValueError("メールアドレスの生成に失敗しました。件数を減らしてください。")
    
    def generate_guardian_name_with_furigana(self, student_last_name, student_last_name_kana):
        """保護者氏名とふりがなを生成（学生と同じ苗字）"""
        first_name, first_name_kana = random.choice(self.guardian_first_names)
        full_name = f"{student_last_name} {first_name}"
        full_name_kana = f"{student_last_name_kana} {first_name_kana}"
        
        return full_name, full_name_kana
    
    def generate_data(self):
        # ファイルパスのチェック
        if not self.file_path.get():
            messagebox.showerror("エラー", "Excelファイルを選択してください")
            return
        
        # 少なくとも1つの項目が選択されているか確認
        if not any([
            self.student_id_enabled.get(),
            self.name_enabled.get(),
            self.furigana_enabled.get(),
            self.gender_enabled.get(),
            self.birthdate_enabled.get(),
            self.address_enabled.get(),
            self.guardian_name_enabled.get(),
            self.guardian_furigana_enabled.get(),
            self.guardian_phone_enabled.get(),
            self.email_enabled.get()
        ]):
            messagebox.showwarning("警告", "少なくとも1つの項目を選択してください")
            return
        
        try:
            count = min(int(self.data_count.get()), 1000)
            start_row = int(self.start_row.get())
            
            # 重複チェック用セットをリセット
            self.used_student_ids.clear()
            self.used_emails.clear()
            self.used_names.clear()
            
            print(f"データ生成開始: {count}件")
            
            # ファイルが存在するか確認
            if os.path.exists(self.file_path.get()):
                wb = openpyxl.load_workbook(self.file_path.get())
            else:
                # 新規ファイルを作成
                wb = Workbook()
                # デフォルトシートを削除
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])
            
            # シートを取得または作成
            if self.sheet_name.get() in wb.sheetnames:
                ws = wb[self.sheet_name.get()]
            else:
                ws = wb.create_sheet(self.sheet_name.get())
            
            # データを生成
            for i in range(count):
                row = start_row + i
                student_name = ""
                student_furigana = ""
                student_last_name = ""
                student_last_name_kana = ""
                gender = None
                
                # 性別を先に決定（性別が有効な場合）
                if self.gender_enabled.get():
                    gender = random.choice(["男", "女"])
                    is_male = (gender == "男")
                else:
                    is_male = None
                
                # 名前とふりがなを生成
                if self.name_enabled.get() or self.furigana_enabled.get():
                    student_name, student_furigana, student_last_name, student_last_name_kana =                         self.generate_name_with_furigana(is_male)
                    
                    if self.name_enabled.get():
                        cell = ws[f"{self.name_column.get().upper()}{row}"]
                        cell.value = student_name
                        print(f"名前生成: {student_name}")
                    
                    if self.furigana_enabled.get():
                        cell = ws[f"{self.furigana_column.get().upper()}{row}"]
                        cell.value = student_furigana
                        print(f"ふりがな生成: {student_furigana}")
                
                # 学生番号を生成
                if self.student_id_enabled.get():
                    student_id = self.generate_student_id()
                    cell = ws[f"{self.student_id_column.get().upper()}{row}"]
                    cell.value = student_id
                    print(f"学生番号生成: {student_id}")
                
                # 性別を記録
                if self.gender_enabled.get():
                    cell = ws[f"{self.gender_column.get().upper()}{row}"]
                    cell.value = gender
                
                # 生年月日を生成
                if self.birthdate_enabled.get():
                    birthdate = self.generate_birthdate()
                    cell = ws[f"{self.birthdate_column.get().upper()}{row}"]
                    cell.value = birthdate
                
                # 住所を生成
                if self.address_enabled.get():
                    address = self.generate_tokyo_address()
                    cell = ws[f"{self.address_column.get().upper()}{row}"]
                    cell.value = address
                
                # 保護者氏名とふりがなを生成
                if self.guardian_name_enabled.get() or self.guardian_furigana_enabled.get():
                    if not student_last_name:
                        temp_name, temp_furigana, student_last_name, student_last_name_kana =                             self.generate_name_with_furigana()
                    
                    guardian_name, guardian_furigana = self.generate_guardian_name_with_furigana(
                        student_last_name, student_last_name_kana
                    )
                    
                    if self.guardian_name_enabled.get():
                        cell = ws[f"{self.guardian_name_column.get().upper()}{row}"]
                        cell.value = guardian_name
                    
                    if self.guardian_furigana_enabled.get():
                        cell = ws[f"{self.guardian_furigana_column.get().upper()}{row}"]
                        cell.value = guardian_furigana
                
                # 保護者電話番号を生成
                if self.guardian_phone_enabled.get():
                    phone = self.generate_phone_number()
                    cell = ws[f"{self.guardian_phone_column.get().upper()}{row}"]
                    cell.value = phone
                
                # メールアドレスを生成
                if self.email_enabled.get():
                    email = self.generate_email()
                    cell = ws[f"{self.email_column.get().upper()}{row}"]
                    cell.value = email
            
            # ファイルを保存
            wb.save(self.file_path.get())
            wb.close()
            
            # 重複率を計算
            duplicate_count = count - len(self.used_names)
            duplicate_rate = (duplicate_count / count) * 100 if count > 0 else 0
            
            print("データ生成完了")
            print(f"重複数: {duplicate_count}件 ({duplicate_rate:.1f}%)")
            
            messagebox.showinfo("完了", f"{count}件のデータ生成が完了しました\n\nファイル: {self.file_path.get()}\n\n名前の重複: {duplicate_count}件 ({duplicate_rate:.1f}%)")
            
        except PermissionError:
            messagebox.showerror("エラー", "ファイルが開かれています。\nExcelファイルを閉じてから再度実行してください。")
        except ValueError as e:
            messagebox.showerror("エラー", f"入力値が不正です: {str(e)}")
        except Exception as e:
            messagebox.showerror("エラー", f"エラーが発生しました: {str(e)}\n\n詳細: {type(e).__name__}")
            import traceback
            traceback.print_exc()

if __name__ == "__main__":
    root = tk.Tk()
    app = DummyDataGenerator(root)
    root.mainloop()
